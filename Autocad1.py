import math
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from termcolor import colored


# Опорный узел, с которого начинается расчет (источник)
Support_node = 'ПП-НО2'
# Падение U и сопротивление до источника
dUfromTP = 0.019
ZfromTP = 0.005


cable_df = pd.DataFrame(np.array([['ЭО_Воздушная линия', 'СИП-2А-3x25+1x35', 25, 35, 13.1, 1.47, 1.05],
                                  ['ЭО_Кабельная линия', 'ПвВГнг(А)-4х16', 16, 16, 21.9, 1.37, 1.37],
                                  ]),
                        columns=['Type', 'NameCable', 'Sf', 'So', 'alfa', 'Rf', 'Ro'])

bracket_df = pd.DataFrame(np.array([['К20-0,5-0,5-0-1',        1,   0],
                                    ['К21(90)-0,5-0,5-0-1',    2,  90],
                                    ['К21(135)-0,5-0,5-0-1',   2, 135],
                                    ['К21-0,5-0,5-0-1',        2, 180],
                                    ['К22-0,5-0,5-0-1',        3,  90]
                                    ]),
                          columns=['Type', 'Count', 'Angle'], dtype=str)
bracket_df.loc[:, ['Count', 'Angle']] = bracket_df.loc[:, ['Count', 'Angle']].astype('float')

Support_df = pd.DataFrame(np.array([['ОГККВ-7,5 С3', 'ФМ-0,325-3,0', 17000, 40000, 7.5],
                                    ['ОГККВ-7,5 С1', 'ФМ-0,325-3,0',  7500, 17000, 7.5],
                                    ['ОГККВ-7,5 У1', 'ФМ-0,219-2,2',     1,  7500, 7.5],
                                    ['ОГКС-7,5',     'ФМ-0,219-2,2',     0,     1, 7.5],
                                    ['ОГКС-4',       'ФМ-0,159-2,0',     0,     1,   4]
                                    ]),
                          columns=['Support_Type', 'Found_Type', 'min_F', 'max_F', 'height'], )
Support_df.loc[:, ['min_F', 'max_F', 'height']] = Support_df.loc[:, ['min_F', 'max_F', 'height']].astype('float')

# Функции для поэтапной обработки, в целом названия говорят сами за себя.
# Честно сказать оформление очень херовенькое, весь скрипт тупо написание отдельных функций и поэтапная обработка
def of_TXT_in_df(file):
    file = file.read().rstrip().split('\n')
    g = []
    for i in file:
        lst = i.split('\t')
        g.append(lst)
    file = pd.DataFrame(g)
    return file


def points_line(df):
    df = df.loc[:, 'COORDS':]
    df.columns = range(df.shape[1])
    for g in range(df.shape[1]):
        for i in range(df.shape[0]):
            if df.loc[i, g] is not None:
                xy = df.loc[i, g].strip('(').strip(')').split()
                df.loc[i, f'POINTx{g}'] = round(float(xy[1]), -3)
                df.loc[i, f'POINTy{g}'] = round(float(xy[2]), -3)
            else:
                df.loc[i, f'POINTx{g}'] = None
                df.loc[i, f'POINTy{g}'] = None
    return df


def points_line1(line):
    df = line.loc[:, 'COORDS':]
    df.columns = range(df.shape[1])
    for index, row in df.iterrows():
        xy1 = df.loc[index, 0].strip('(').strip(')').split()
        xy2 = df.loc[index, len(row.dropna()) - 1].strip('(').strip(')').split()
        line.loc[index, 'POINTx1'] = round(float(xy1[1]), -3)
        line.loc[index, 'POINTy1'] = round(float(xy1[2]), -3)
        line.loc[index, 'POINTx2'] = round(float(xy2[1]), -3)
        line.loc[index, 'POINTy2'] = round(float(xy2[2]), -3)
    return line


def points(df):
    for i in range(len(df)):
        xy = df.loc[i, 'POINT'].strip('(').strip(')').split()
        df.loc[i, 'POINTx'] = round(float(xy[0]), -3)
        df.loc[i, 'POINTy'] = round(float(xy[1]), -3)
    return df


def get_distance(unit1, unit2):
    phi = abs(unit2 - unit1) % 360
    sign = 1
    # used to calculate sign
    if not ((0 <= unit1 - unit2 <= 180) or
            (-180 >= unit1 - unit2 >= -360)):
        sign = -1
    if phi > 180:
        result = 360 - phi
    else:
        result = phi
    return abs(int(round(result * sign / 5.0) * 5.0))


def light_in_support(Support, light):
    print('____________________________________________________________________________________')
    print(colored('              Начало привязки светильников к опорам!', 'yellow'))
    number_light = 0
    power_light = 0
    for index, row in Support.iterrows():
        ss = light[(light.POINTx == row.POINTx) & (light.POINTy == row.POINTy)]
        power = ss.loc[:, ['POWER', 'Angle1']].sort_values(by=['POWER'], ascending=False)
        Support.loc[index, 'Count'] = int(len(power))
        Support.loc[index, 'Power'] = power.POWER.sum()
        if len(power) < 2:
            Support.loc[index, 'Angle'] = 0
        elif len(power) == 2:
            Support.loc[index, 'Angle'] = get_distance(power.Angle1[power.index[0]], power.Angle1[power.index[1]])
        elif len(power) >= 3:
            angle1 = get_distance(power.Angle1[power.index[0]], power.Angle1[power.index[1]])
            angle2 = get_distance(power.Angle1[power.index[0]], power.Angle1[power.index[2]])
            Support.loc[index, 'Angle'] = min(angle1, angle2)
        light_lst = str('Волна Мини LED ')
        for p in range(len(power)):
            Support.loc[index, f'light{p + 1}'] = power.POWER[power.index[p]]
            light.loc[power.index[p], 'POWER1'] = power.POWER[power.index[p]]
            power_light = power_light + power.POWER[power.index[p]]
            number_light = number_light + 1
            if len(light_lst) > 15:
                light_lst = light_lst +str(',')
            light_lst = light_lst + str(int(power.POWER[power.index[p]]))
        light_lst = light_lst + 'Вт'
        Support.loc[index, 'LIGHT'] = light_lst
    light1 = light.isnull()
    light1 = light1[light1.POWER1 == True]
    light = light.loc[light1.index, ['POINTx', 'POINTy', 'POWER']]
    print(light)
    print(colored(f'Количество светильников {number_light}', 'yellow'))
    print(colored(f'Мощность всех светильников {power_light}', 'yellow'))
    print(colored('              Конец привязки светильников к опорам!', 'yellow'))
    return Support


def cable_in_support(line, Support):
    print('____________________________________________________________________________________')
    print(colored('              Начало привязки кабелей к опорам ', 'magenta'))
    for index, row in line.iterrows():
        try:
            a = Support[(Support.POINTx == row.POINTx1) & (Support.POINTy == row.POINTy1)].N_PYLON
            b = Support[(Support.POINTx == row.POINTx2) & (Support.POINTy == row.POINTy2)].N_PYLON
            line.loc[index, 'N_PYLON1'] = str(a[a.index[0]])
            line.loc[index, 'N_PYLON2'] = str(b[b.index[0]])
        except:
            print('Не найдена опора в линии между ', line.loc[index, 'N_PYLON1'], line.loc[index, 'N_PYLON2'])
        # print('Привязка линии длиной', line.loc[index, 'LENGTH'], 'слоя', line.loc[index, 'LAYER'], 'между',
        #       line.loc[index, 'N_PYLON1'], line.loc[index, 'N_PYLON2'])
    line['LENGTH'] = line['LENGTH'].astype('float')
    print(colored('              Конец привязки кабелей к опорам ', 'magenta'))
    return line


def draw_graph(G):
    plt.figure()
    pos = nx.kamada_kawai_layout(G)
    edges = G.edges()
    colors = [G[u][v][0]['LAYER'] for u, v in edges]
    color = []
    for i in colors:
        if i == 'ЭО_Кабельная линия':
            color.append('g')
        if i == 'ЭО_Воздушная линия':
            color.append('r')
    nx.draw(G, pos, edge_color=color, with_labels=True)
    plt.show(block=False)
    # print(nx.dfs_successors(G,"N"))
    # print(nx.dfs_predecessors(G,"N"))
    # print(G["N"])


def draw_graph1(Ga, Support):
    Support = Support.set_index('N_PYLON')
    plt.figure()
    pos = nx.kamada_kawai_layout(Ga)
    nx.draw_networkx_nodes(Ga, pos, node_color='b', node_size=150, alpha=0.2, label=True)
    nx.draw_networkx_labels(Ga, pos, font_size=10)
    ax = plt.gca()
    for e in Ga.edges:
        color = nx.get_edge_attributes(Ga, 'LAYER')
        if color[e] == 'ЭО_Кабельная линия':
            color = 'g'
        elif color[e] == 'ЭО_Воздушная линия':
            color = 'r'
        ax.annotate("",
                    xy=pos[e[0]], xycoords='data',
                    xytext=pos[e[1]], textcoords='data',
                    arrowprops=dict(arrowstyle="->", color=color,
                                    shrinkA=5, shrinkB=5,
                                    patchA=None, patchB=None,
                                    connectionstyle="arc3,rad=rrr".replace('rrr', str(0.2 * e[2])
                                                                           ),
                                    ),
                    )
        ax.text(pos[e[1]][0] - 0.03, pos[e[1]][1] - 0.04, Support.loc[e[1], 'GROUP'],
                fontsize=7)
    plt.axis('off')
    plt.show()
    plt.show()


def support_selection(G, Support):
    print('____________________________________________________________________________________')
    print(colored('              Начало подбора опор кронштейнов и фундаментов', 'blue'))

    '''В данной функции происходит векторное сложение всех воздушных линий поключенных к опоре,
    Из графа выгружаются все воздушные линии опоры и после происходит сложение их половинок.
    При этом образуется coefficient который характиризует натяжение опоры.
    Сделал так просто потому что слишком долго изучать как правильно считается нагрузка,
    и надо побыстрее доделать прототип'''
    Support = Support.set_index('N_PYLON')
    for N_PYLON in Support.index:
        weight_x, weight_y = 0, 0
        coefficient = 0
        x1 = int(Support[Support.index == N_PYLON].POINTx)
        y1 = int(Support[Support.index == N_PYLON].POINTy)
        in_out = []
        for CN_PYLON in G[N_PYLON]:
            for num in G[N_PYLON][CN_PYLON]:
                in_out.append(G[N_PYLON][CN_PYLON][num]['LAYER'])
                if G[N_PYLON][CN_PYLON][num]['LAYER'] == 'ЭО_Воздушная линия':
                    x2 = int(Support[Support.index == CN_PYLON].POINTx)
                    y2 = int(Support[Support.index == CN_PYLON].POINTy)
                    delt_x = (x2 - x1) / 2
                    delt_y = (y2 - y1) / 2
                    weight_x = weight_x + delt_x
                    weight_y = weight_y + delt_y
        coefficient = math.sqrt((weight_x ** 2) + (weight_y ** 2))
        '''Следующее условие необходимо для того чтобы воздушная опора с натяжением 0 
        не учитывалась как кабельная.Поэтому ему просто присваивается 7.
        Знаю что тупо но зато очень просто и лишает многих проблем. 
         Думал создать какой нибудь параметр который за это отвечает,
        но так мне кажется проще'''
        if coefficient == 0 and ('ЭО_Воздушная линия' in in_out):
            coefficient = 7
        Support.loc[N_PYLON, 'PYLON'] = coefficient
        '''Собственно сами условия выбора типа опоры, они очень условные и ни на что не влияют.
        Сделано скорее для наглядности и проверки'''
        if coefficient == 0:
            Support.loc[N_PYLON, 'PYLON_Type'] = "Обычная"
        elif 0 < coefficient < 7500:
            Support.loc[N_PYLON, 'PYLON_Type'] = "Проходная"
        elif 7500 <= coefficient < 17000:
            Support.loc[N_PYLON, 'PYLON_Type'] = "Концевая"
        elif 17000 <= coefficient:
            Support.loc[N_PYLON, 'PYLON_Type'] = "Угловая"
    Support = Support.reset_index()
    '''Вторая часть в которой уже конкретно подбираются кронштейны опоры и фундаменты.
    Надо бы ее соединить с верхней частью кода, но как нибудь потом'''
    for index, row in Support.iterrows():
        try:
            Support.loc[index, 'BRACKET'] = bracket_df[(bracket_df.Count == Support.loc[index, 'Count']) &
                                                       (bracket_df.Angle == Support.loc[index, 'Angle'])].Type.values
        except:
            print('не найден кронштейн для опоры', Support.loc[index, 'N_PYLON'], 'c', Support.loc[index, 'Count'],
                  'светильником(ами) и углом', Support.loc[index, 'Angle'])
        try:
            a = Support_df[(Support_df.min_F <= row.PYLON) &
                           (Support_df.max_F > row.PYLON) & (row.HEIGHT == Support_df.height)].loc[:,
                ['Support_Type', 'Found_Type']]
            Support.loc[index, 'SUPPORT_TYPE'] = a.Support_Type.values
            Support.loc[index, 'FOUND_TYPE'] = a.Found_Type.values
        except:
            print('не найдена опора для ', Support.loc[index, 'N_PYLON'])
    print(colored('              Конец подбора опор кронштейнов и фундаментов', 'blue'))
    return Support


def support_set_power(G, Support):
    Support = Support.set_index('N_PYLON')
    tree = nx.dfs_successors(G, Support_node)
    for GROUP in Support['GROUP'].unique():
        Support[f'{GROUP}'] = Support[Support.GROUP == GROUP].Power
        Support[f'{GROUP}'] = Support[f'{GROUP}'].fillna(0)
        for i in reversed(list(tree.keys())):
            for g in tree[i]:
                Support.loc[i, f'{GROUP}'] = Support.loc[i, f'{GROUP}'] + Support.loc[g, f'{GROUP}']
    Support = Support.reset_index()
    return Support


def cable_list(G, Support, Support_node):
    """ Создается df из правильно направленных ребер графа и по порядку алгоритма поиска в глубину """
    Support = Support.set_index('N_PYLON')
    cable_list = list(nx.dfs_edges(G, Support_node))
    cable_list = pd.DataFrame(data=cable_list)
    cable_list = cable_list.rename(columns={0: 'N_PYLON1', 1: 'N_PYLON2'})
    ''' Присваиваются параметры из  df по линиям(Line), 
    так же во втором цикле for присваиваются все мощности по группам которые были в таблице Support'''
    for index, row in cable_list.iterrows():
        Line_row = Line[((Line.N_PYLON1 == row['N_PYLON1']) & (Line.N_PYLON2 == row['N_PYLON2'])) |
                        ((Line.N_PYLON1 == row['N_PYLON2']) & (Line.N_PYLON2 == row['N_PYLON1']))]
        cable_list.loc[index, 'LENGTH'] = Line_row.LENGTH.values
        cable_list.loc[index, 'LAYER'] = Line_row.LAYER.values
        cable_list.loc[index, '//HANDLE'] = Line_row['//HANDLE'].values
        for GROUP in Support['GROUP'].unique():
            cable_list.loc[index, f'{GROUP}'] = Support.loc[row['N_PYLON2'], GROUP]
    '''Здесь добавляется по высоте опоры к кабельной линиии для того чтобы учесть переход земля воздух
    Первым условием создаются два списка из типа приходящих и отходящих линий.
    Вторым условием они  сравнениваются  и присваиваются нужные значения  '''
    cable_list['ADD'] = float(1)
    for index, row in cable_list.iterrows():
        if row.LAYER == 'ЭО_Кабельная линия':
            a = list(cable_list[cable_list.N_PYLON2 == row.N_PYLON1].LAYER.unique())
            b = list(cable_list[cable_list.N_PYLON1 == row.N_PYLON2].LAYER.unique())
            if a != [row.LAYER] and len(a) > 0:
                cable_list.loc[index, 'ADD'] = cable_list.loc[index, 'ADD'] + float(
                    Support.loc[row['N_PYLON1'], 'HEIGHT'])
            if b != [row.LAYER] and len(b) > 0:
                cable_list.loc[index, 'ADD'] = cable_list.loc[index, 'ADD'] + float(
                    Support.loc[row['N_PYLON2'], 'HEIGHT'])

    ''' Из одного кабельного журнала по всем группам формируется несколько для каждой группы и 
    и соединяются в один '''
    cable_list_new = pd.DataFrame()
    for GROUP in Support['GROUP'].unique():
        df = cable_list[cable_list[f'{GROUP}'] != 0].loc[:,
             ('//HANDLE', 'N_PYLON1', 'N_PYLON2', 'LENGTH', 'LAYER', 'ADD', f'{GROUP}')]
        df = df.rename(columns={f'{GROUP}': 'POWER'})
        df['GROUP'] = GROUP
        cable_list_new = pd.concat([cable_list_new, df])
    return cable_list_new


def calculation_of_voltage_drop_and_short_circuit_currents(dU_Tkz):
    ####################### РАСЧЕТ ПАДЕНИЯ НАПРЯЖЕНИЯ И ТКЗ ##############################################
    """ Расчет dU и Z_area для каждого участка в зависимости от типа кабеля """
    dU_Tkz = dU_Tkz.reset_index(drop=True)
    dU_Tkz.LENGTH = dU_Tkz.LENGTH / 1000
    for index, row in dU_Tkz.iterrows():
        dU_Tkz.loc[index, 'dU_area'] = float(cable_df[cable_df['Type'] == row['LAYER']].alfa) * \
                                       row['POWER'] / 1000 * row['LENGTH'] / (10 ** 3) / \
                                       float(cable_df[cable_df['Type'] == row['LAYER']].Sf)
        dU_Tkz.loc[index, 'Z_area'] = (float(cable_df[cable_df['Type'] == row['LAYER']].Ro) +
                                       float(cable_df[cable_df['Type'] == row['LAYER']].Rf)) * \
                                      row['LENGTH'] / 1000
    ''' Переписываются значения в еще один столбик для расчета общих параметров '''
    dU_Tkz.loc[:, 'dU'] = dU_Tkz.loc[:, 'dU_area']
    dU_Tkz.loc[:, 'Z'] = dU_Tkz.loc[:, 'Z_area']
    ''' Добавление первоначальных значений dU и Z до первого узла(расчитываются вручную) '''
    for index, row in dU_Tkz[(dU_Tkz.N_PYLON1 == Support_node)].iterrows():
        dU_Tkz.loc[index, 'dU'] = dU_Tkz.loc[index, 'dU'] + dUfromTP
        dU_Tkz.loc[index, 'Z'] = dU_Tkz.loc[index, 'Z'] + ZfromTP
    ''' Создания списка ребер по алгоритму поиска в глубину от опорного узла
     Последовательный поиск ребер из кабельного журнала и последовательное сложение dU и dZ
     И присвоение получившехся параметров '''
    tree = list(nx.dfs_edges(G, Support_node))
    for gr in dU_Tkz['GROUP'].unique():
        for su in tree:
            cable_next = dU_Tkz[(dU_Tkz.N_PYLON1 == su[1]) &
                                (dU_Tkz.GROUP == gr)]
            cable_now = dU_Tkz[(dU_Tkz.N_PYLON1 == su[0]) &
                               (dU_Tkz.N_PYLON2 == su[1]) &
                               (dU_Tkz.GROUP == gr)]
            if len(cable_next.index.values) > 0:
                for index, row in cable_next.iterrows():
                    dU_Tkz.loc[index, 'dU'] = dU_Tkz.loc[index, 'dU'] + cable_now.dU.values
                    dU_Tkz.loc[index, 'Z'] = dU_Tkz.loc[index, 'Z'] + cable_now.Z.values
    ''' расчет ТКЗ '''
    dU_Tkz.loc[:, 'Ikz'] = 0.22 / dU_Tkz.loc[:, 'Z'] * 1000
    ''' Замена слоя на кабель '''
    for index, row in dU_Tkz.iterrows():
        dU_Tkz.loc[index, 'LAYER'] = cable_df[cable_df['Type'] == row['LAYER']].NameCable.values
    ''' Сортировка по группам в порядке возрастания групп и присвоение индекса '''
    dU_Tkz = dU_Tkz.reset_index()
    dU_Tkz = dU_Tkz.sort_values(by=['GROUP', 'index'])
    dU_Tkz = dU_Tkz.reset_index(drop=True)
    dU_Tkz = dU_Tkz.drop(columns='index')
    ''' Округление'''
    dU_Tkz.loc[:, 'dU_area':'Ikz'] = dU_Tkz.loc[:, 'dU_area':'Ikz'].round(4)
    dU_Tkz.loc[:, 'LENGTH'] = dU_Tkz.loc[:, 'LENGTH'].round(1)
    ''' Формирование кабельного журнала '''
    cable_list = dU_Tkz.loc[:, ['//HANDLE', 'GROUP', 'N_PYLON1', 'N_PYLON2', 'LAYER', 'LENGTH', 'ADD']]
    cable_ground = cable_df[cable_df.Type == 'ЭО_Кабельная линия'].NameCable.iloc[0]
    cable_air = cable_df[cable_df.Type == 'ЭО_Воздушная линия'].NameCable.iloc[0]

    cable_list['GROUND'] = cable_list[cable_list.LAYER == cable_ground].LENGTH
    cable_list['AIR'] = cable_list[cable_list.LAYER == cable_air].LENGTH
    cable_list['GROUND'] = cable_list['GROUND'].fillna(0)
    cable_list['AIR'] = cable_list['AIR'].fillna(0)
    cable_list['SUMM'] = cable_list['GROUND'] + cable_list['AIR'] + cable_list['ADD']

    return cable_list, dU_Tkz


def crossing(line, Tube, cable_list):
    print('____________________________________________________________________________________')
    print(colored('              Начало привязки пересечек к линиям', 'grey'))
    line = line.set_index('//HANDLE')
    line = line.loc[:, 'COORDS':]
    line.columns = range(line.shape[1])
    print('          Таблица дубликатов уникальных номеров пересечек')
    print(Tube[Tube.duplicated(subset=['N_CROSSING'], keep=False)])
    cable_list.loc[:, 'CROSSING'] = ''
    cable_list['CROSSING'] = cable_list['CROSSING'].astype('str')
    unique_cross = list(Tube.N_CROSSING.sort_values().unique())
    print('Название уникальных пересечек:', unique_cross)
    duplicate_cross = []

    for unique_tube in Tube.WIDTH.unique():
        cable_list[f'TUBE{unique_tube}'] = 0
    for itube, rowtube in Tube.iterrows():
        xy = rowtube.POINT.strip('(').strip(')').split()
        x = float(xy[0])
        y = float(xy[1])
        for iline, rowline in line.iterrows():
            for i in range(len(rowline.dropna()) - 1):
                x1y1 = rowline[i].strip('(').strip(')').split()
                x1 = float(x1y1[1])
                y1 = float(x1y1[2])
                x2y2 = rowline[i + 1].strip('(').strip(')').split()
                x2 = float(x2y2[1])
                y2 = float(x2y2[2])
                col = round(float(((x - x1) * (y2 - y1)) - ((y - y1) * (x2 - x1))), -7)
                if (col == 0) & ((x1 <= x <= x2) | (x2 <= x <= x1)) \
                        & ((y1 <= y <= y2) | (y2 <= y <= y1)):
                    N_PYLON1 = cable_list[cable_list['//HANDLE'] == iline].N_PYLON1.values[0]
                    N_PYLON2 = cable_list[cable_list['//HANDLE'] == iline].N_PYLON2.values[0]
                    number = len(cable_list[cable_list['//HANDLE'] == iline].loc[:, ['N_PYLON1', 'N_PYLON2']])
                    print('Для пересечки номер ', rowtube.N_CROSSING,
                          'найдено', number, 'кабеля между', N_PYLON1, N_PYLON2)
                    index = cable_list[cable_list['//HANDLE'] == iline].index
                    cable_list.loc[index, 'CROSSING'] = \
                        cable_list.loc[index, 'CROSSING'] + '(' + rowtube.N_CROSSING + ')'
                    cable_list.loc[index, f'TUBE{rowtube.WIDTH}'] = \
                        cable_list.loc[index, f'TUBE{rowtube.WIDTH}'] + float(rowtube.LENGTH) / 1000
                    unique_cross.remove(rowtube.N_CROSSING)
    print('Пересечки не привязанные к кабелям', unique_cross)
    print(colored('              Конец привязки пересечек к линиям', 'grey'))
    return cable_list


def dU_Tkz_decor(dU_Tkz):
    dU_Tkz_of = dU_Tkz.loc[:,
                ['GROUP', 'N_PYLON1', 'N_PYLON2', 'LENGTH', 'POWER', 'LAYER', 'dU_area', 'dU', 'Z_area', 'Z', 'Ikz']]
    dU_Tkz_of = dU_Tkz_of.rename(columns={'N_PYLON1': 'Начало\n участка',
                                          'N_PYLON2': 'Конец\n участка',
                                          'LENGTH': 'Длина\nучастка,м',
                                          'LAYER': 'Марка провода',
                                          'POWER': 'Рр на\nучастке,\n Вт',
                                          'GROUP': 'Группа',
                                          'dU_area': '∆U% на\nуч.,В',
                                          'Z_area': 'Z петли\nучастка, Ом',
                                          'dU': '∆U%,\nот ТП,В',
                                          'Z': 'Z петли\nдо ТП, Ом',
                                          'Ikz': 'Iкз 1,\n А', })
    dU_Tkz_of.to_excel('dU и ТКЗ.xlsx', sheet_name='0')
    wb = openpyxl.load_workbook('dU и ТКЗ.xlsx')
    ws = wb.active
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['j'].width = 9
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 9

    # Выравнивание
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          wrap_text=True)

    # Шрифт
    font = Font(name='Calibri', italic=True, )

    # Заливка четных групп
    patternfill1 = PatternFill(fill_type='solid',
                               start_color='00CCFFFF',
                               end_color='00CC99FF')
    # Заливка нечетных групп
    patternfill2 = PatternFill(fill_type='solid',
                               start_color='00FFCC99',
                               end_color='00FFCC99')
    # Заливка  красным
    patternfill3 = PatternFill(fill_type='solid',
                               start_color='00FF7171',
                               end_color='00FF7171')

    # Граница №1 для обычных строк
    border1 = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))

    # Граница №2 для конечных строк
    border2 = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thick', color='FF000000'))

    col = ws['B1':'L1']
    for cells in col:
        for cell in cells:
            cell.alignment = alignment
            cell.font = font

    gr = str()
    score = 0
    for i in range(2, ws.max_row + 1):
        if gr != ws.cell(row=i, column=2).value:
            gr = ws.cell(row=i, column=2).value
            score += 1
        if ws.cell(row=i, column=12).value == dU_Tkz.Ikz.min():
            patternfill = patternfill3
        elif score % 2 == 1:
            patternfill = patternfill1
        elif score % 2 != 1:
            patternfill = patternfill2
        col = ws[f'B{i}':f'L{i}']
        if ws.cell(row=i, column=4).value == ws.cell(row=i + 1, column=3).value:
            for cells in col:
                for cell in cells:
                    cell.alignment = alignment
                    cell.font = font
                    cell.fill = patternfill
                    cell.border = border1
        else:
            for cells in col:
                for cell in cells:
                    cell.alignment = alignment
                    cell.font = font
                    cell.fill = patternfill
                    cell.border = border2
    wb.save("dU и ТКЗ.xlsx")


def cable_list_decor(cable_list):
    list_name = ['GROUP', 'N_PYLON1', 'N_PYLON2', 'LAYER', 'LENGTH', 'AIR', 'GROUND', 'ADD', 'SUMM']
    dict_name = {'GROUP': 'Группа',
                 'N_PYLON1': 'Начало',
                 'N_PYLON2': 'Конец',
                 'LAYER': 'Марка провода',
                 'LENGTH': 'Длина ,м',
                 'POWER': 'Рр на\nучастке,\n Вт',
                 'AIR': 'По воздуху',
                 'GROUND': 'В траншеи\n (в гофр. тр.D50мм)',
                 'ADD': 'В щите/опоре.\nНа разделку/провис.',
                 'SUMM': 'Итого кабеля'}

    if blocks[(blocks.loc[:, 1] == 'Труба с размером')].shape[0] > 0:
        list_name.append('CROSSING')
        dict_name.update({'CROSSING': 'Номер\nпересечения'})
        for i in tube.WIDTH.unique():
            list_name.append(f'TUBE{i}')
            dict_name.update({f'TUBE{i}': f'В траншеи\n(в ПНД D{i}мм)'})

    cable_list_of = cable_list.loc[:, list_name]
    cable_list_of = cable_list_of.rename(columns=dict_name)

    cable_list_of.to_excel('Кабельный журнал.xlsx', sheet_name='0')
    wb = openpyxl.load_workbook('Кабельный журнал.xlsx')
    ws = wb.active

    # Выравнивание и разворачивание первой строки
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=90,
                          wrap_text=True)
    # Выравнивание
    alignment1 = Alignment(horizontal='center',
                           vertical='center',
                           wrap_text=True)

    # Шрифт
    font1 = Font(name='Calibri', italic=True, )

    # Шрифт жирный
    font2 = Font(name='Calibri', italic=True, bold=True, )

    # Заливка четных групп  кабельная линия
    patternfill_1gr = PatternFill(fill_type='solid',
                                  start_color='00CCFFFF',
                                  end_color='00CCFFFF')
    # Заливка четных групп воздушная линия
    patternfill_1air = PatternFill(fill_type='solid',
                                   start_color='00E6FFFF',
                                   end_color='00E6FFFF')
    # Заливка нечетных групп  кабельная линия
    patternfill_2gr = PatternFill(fill_type='solid',
                                  start_color='00FFCC99',
                                  end_color='00FFCC99')
    # Заливка нечетных групп воздушная линия
    patternfill_2air = PatternFill(fill_type='solid',
                                   start_color='00FFE6CD',
                                   end_color='00FFE6CD')

    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value and cell.row > 1:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4

    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                cell.alignment = alignment
                cell.font = font1
                cell.border = border

    gr = str()
    score = 0
    for i in range(2, ws.max_row + 1):
        if gr != ws.cell(row=i, column=2).value:
            gr = ws.cell(row=i, column=2).value
            score += 1
        if ws.cell(row=i, column=5).value == cable_df[cable_df.Type == 'ЭО_Воздушная линия'].NameCable.values:
            if score % 2 == 1:
                patternfill = patternfill_1air
            elif score % 2 != 1:
                patternfill = patternfill_2air
        elif ws.cell(row=i, column=5).value == cable_df[cable_df.Type == 'ЭО_Кабельная линия'].NameCable.values:
            if score % 2 == 1:
                patternfill = patternfill_1gr
            elif score % 2 != 1:
                patternfill = patternfill_2gr
        row = ws[i]
        for cell in row:
            if cell.column > 1:
                if cell.column == 10:
                    cell.font = font2
                else:
                    cell.font = font1
                while cell.column < 11:
                    cell.alignment = alignment1
                    break
                cell.fill = patternfill
                cell.border = border
                if cell.value == 0:
                    cell.value = ''

    wb.save("Кабельный журнал.xlsx")


def Support_decor(Support):
    Support = Support.drop(Support[Support.N_PYLON == Support_node].index)
    Support = Support.sort_values(by=['N_PYLON'])
    Support = Support.reset_index(drop=True)
    list_name = ['N_PYLON', 'GROUP', 'SUPPORT_TYPE', 'BRACKET', 'FOUND_TYPE', 'LIGHT']
    dict_name = {'GROUP': 'Группа',
                 'N_PYLON': '№ Опоры',
                 'SUPPORT_TYPE': 'Тип опоры',
                 'BRACKET': 'Кронштейн',
                 'FOUND_TYPE': 'Закладная\nдеталь',
                 'LIGHT': 'Светильники',}
    Support = Support.loc[:, list_name]
    Support = Support.rename(columns=dict_name)

    Support.to_excel('Ведомость.xlsx', sheet_name='0')
    wb = openpyxl.load_workbook('Ведомость.xlsx')
    ws = wb.active


    # Выравнивание
    alignment1 = Alignment(horizontal='center',
                           vertical='center',
                           wrap_text=True)

    alignment2 = Alignment(horizontal='left',
                           vertical='center',
                           wrap_text=True)

    # Шрифт
    font1 = Font(name='Calibri', italic=True, )

    # Шрифт жирный
    font2 = Font(name='Calibri', italic=True, bold=True, )

    # Заливка четных групп  кабельная линия
    patternfill_1gr = PatternFill(fill_type='solid',
                                  start_color='00CCFFFF',
                                  end_color='00CCFFFF')
    # Заливка четных групп воздушная линия
    patternfill_1air = PatternFill(fill_type='solid',
                                   start_color='00E6FFFF',
                                   end_color='00E6FFFF')
    # Заливка нечетных групп  кабельная линия
    patternfill_2gr = PatternFill(fill_type='solid',
                                  start_color='00FFCC99',
                                  end_color='00FFCC99')
    # Заливка нечетных групп воздушная линия
    patternfill_2air = PatternFill(fill_type='solid',
                                   start_color='00FFE6CD',
                                   end_color='00FFE6CD')

    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value and cell.row > 1:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4

        for row in ws.rows:
            for cell in row:
                if cell.row == 1:
                    cell.alignment = alignment1
                    cell.font = font2
                    cell.border = border

                if cell.row > 1:
                    if cell.column > 1:
                        cell.alignment = alignment1
                        cell.font = font1
                        cell.border = border
                        if cell.column == 7:
                            cell.alignment = alignment2
    wb.save('Ведомость.xlsx')


# Создание двух df по выгруженным тхт файлам
blocks = open('OUT_block.txt')
blocks = of_TXT_in_df(blocks)
line = open('OUT_line.txt')
line = of_TXT_in_df(line)

# Обработка df, разделение на df Support,light,Line
# Line - кабельный журнал
# Support - Ведомость опор
INDEXblocks = blocks[(blocks[0] == '//HANDLE')].index
Support = blocks[(blocks.loc[:, 1] == 'опора_промежуточная 0.4')].reset_index(drop=True)
Support.columns = blocks.loc[INDEXblocks[0]]
Support = Support.loc[:, ['NAME_BLOCK', 'POINT', 'N_PYLON', 'GROUP', 'HEIGHT']]
Support['HEIGHT'] = Support['HEIGHT'].astype('float')

# создание df по светильникам
light = blocks[(blocks.loc[:, 1] == 'ЭО_Опора_1_свет')].reset_index(drop=True)
light.columns = blocks.loc[INDEXblocks[1]]
light = light.loc[:, ['NAME_BLOCK', 'POINT', 'POWER', 'Angle1']]

# создание df по пересечкам
if blocks[(blocks.loc[:, 1] == 'Труба с размером')].shape[0] > 0:
    tube = blocks[(blocks.loc[:, 1] == 'Труба с размером')].reset_index(drop=True)
    tube.columns = blocks.loc[INDEXblocks[2]]
    tube = tube.loc[:, ['NAME_BLOCK', 'POINT', 'WIDTH', 'LENGTH', 'N_CROSSING']]

# создание df по линиям
INDEXline = line[(line[0] == '//HANDLE')].index
line.columns = line.loc[INDEXline[0]]
line = line.loc[1:, :]
Line = line[(line['LAYER'] == 'ЭО_Воздушная линия') | (line['LAYER'] == 'ЭО_Кабельная линия')].reset_index(drop=True)

light = points(light)
light = light.assign(POWER=light.POWER.str.replace(r'Вт$', ''))
light.loc[:, ['POWER', 'Angle1']] = light.loc[:, ['POWER', 'Angle1']].astype('float')
light.Angle1 = light.Angle1 / math.pi * 180

# Началась обработка(форматирование координат)
Support = points(Support)
Line = points_line1(Line)
Support.reset_index.__doc__
Support = light_in_support(Support, light)
Line = cable_in_support(Line, Support)

# Формирование первого графа
G = nx.from_pandas_edgelist(Line, 'N_PYLON1', 'N_PYLON2', ['LENGTH', 'LAYER'], create_using=nx.MultiGraph)
# Разделение мощности на опорах  по группам
Support = support_set_power(G, Support)
# Формирование  кабельного листа на основе списка ребер алгоритма
# поиска в грубину.
cable_list = cable_list(G, Support, Support_node)

# Формирование второго мультиграфа в котором учтены группы
G = nx.from_pandas_edgelist(cable_list, 'N_PYLON1', 'N_PYLON2', ['LENGTH', 'LAYER'], create_using=nx.MultiGraph)
# Расчет векторного натяжения опоры с учетом дублирование кабелей разных групп
Support = support_selection(G, Support)

cable_list, dU_Tkz = calculation_of_voltage_drop_and_short_circuit_currents(cable_list)

if blocks[(blocks.loc[:, 1] == 'Труба с размером')].shape[0] > 0:
    cable_list = crossing(line, tube, cable_list)

draw_graph1(G, Support)

dU_Tkz_decor(dU_Tkz)
cable_list_decor(cable_list)
Support_decor(Support)

print(colored('ПОЗДРАВЛЯЮ БЫЛО СЪЭКОНОМЛЕНО НЕМНОГО ЖИЗНИ, после такой работы можно и отдохнуть', 'green'))


